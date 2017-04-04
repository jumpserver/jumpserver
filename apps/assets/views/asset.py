# coding:utf-8
from __future__ import absolute_import, unicode_literals
import json
import uuid

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import load_workbook
from django.conf import settings
from django.db import IntegrityError
from django.urls import reverse
from django.views.generic import TemplateView, ListView, View
from django.views.generic.edit import CreateView, DeleteView, FormView, UpdateView
from django.urls import reverse_lazy
from django.views.generic.detail import DetailView, SingleObjectMixin
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.utils.decorators import method_decorator
from django.core.cache import cache
from django.utils import timezone
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import get_object_or_404, redirect

from common.mixins import JSONResponseMixin
from common.utils import get_object_or_none
from .. import forms
from ..models import Asset, AssetGroup, AdminUser, IDC, SystemUser
from ..hands import AdminUserRequiredMixin
from ..tasks import update_assets_hardware_info


__all__ = ['AssetListView', 'AssetCreateView', 'AssetUpdateView',
           'UserAssetListView', 'AssetModalCreateView', 'AssetDetailView',
           'AssetModalListView', 'AssetDeleteView', 'AssetExportView',
           'BulkImportAssetView', 'AssetRefreshHardwareView',
           ]


class AssetListView(AdminUserRequiredMixin, TemplateView):
    template_name = 'assets/asset_list.html'

    def get_context_data(self, **kwargs):
        context = {
            'app': 'Assets',
            'action': 'asset list',
            'groups': AssetGroup.objects.all(),
            'system_users': SystemUser.objects.all(),
        }
        kwargs.update(context)
        return super(AssetListView, self).get_context_data(**kwargs)


class UserAssetListView(LoginRequiredMixin, TemplateView):
    template_name = 'assets/user_asset_list.html'

    def get_context_data(self, **kwargs):
        context = {
            'app': 'Assets',
            'action': 'asset list',
            'system_users': SystemUser.objects.all(),
        }
        kwargs.update(context)
        return super(UserAssetListView, self).get_context_data(**kwargs)


class AssetCreateView(AdminUserRequiredMixin, CreateView):
    model = Asset
    form_class = forms.AssetCreateForm
    template_name = 'assets/asset_create.html'
    success_url = reverse_lazy('assets:asset-list')

    def form_valid(self, form):
        self.asset = asset = form.save()
        asset.created_by = self.request.user.username or 'Admin'
        asset.date_created = timezone.now()
        asset.save()
        return super(AssetCreateView, self).form_valid(form)

    def get_context_data(self, **kwargs):
        context = {
            'app': 'Assets',
            'action': 'Create asset',
        }
        kwargs.update(context)
        return super(AssetCreateView, self).get_context_data(**kwargs)

    def get_success_url(self):
        update_assets_hardware_info.delay([self.asset])
        return super(AssetCreateView, self).get_success_url()


class AssetModalCreateView(AdminUserRequiredMixin, ListView):
    model = Asset
    form_class = forms.AssetCreateForm
    template_name = 'assets/asset_modal_update.html'
    success_url = reverse_lazy('assets:asset-list')

    def get_queryset(self):
        self.queryset = super(AssetModalCreateView,self).get_queryset()
        self.s = self.request.GET.get('plain_id_lists')
        if "," in str(self.s):
            self.plain_id_lists = [int(x) for x in self.s.split(',')]
        else:
            self.plain_id_lists = [self.s]
        return self.queryset

    def get_context_data(self, **kwargs):
        asset_on_list = Asset.objects.filter(id__in = self.plain_id_lists)
        context = {
            'app': 'Assets',
            'action': 'Bulk Update asset',
            'assets_on_list': asset_on_list,
            'assets_count': len(asset_on_list),
            'plain_id_lists':self.s,
        }
        kwargs.update(context)
        return super(AssetModalCreateView, self).get_context_data(**kwargs)


class AssetUpdateView(AdminUserRequiredMixin, UpdateView):
    model = Asset
    form_class = forms.AssetUpdateForm
    template_name = 'assets/asset_update.html'
    success_url = reverse_lazy('assets:asset-list')

    def get_context_data(self, **kwargs):
        context = {
            'app': 'Assets',
            'action': 'Update asset',
        }
        kwargs.update(context)
        return super(AssetUpdateView, self).get_context_data(**kwargs)

    def form_invalid(self, form):
        print(form.errors)
        return super(AssetUpdateView, self).form_invalid(form)


class AssetDeleteView(AdminUserRequiredMixin, DeleteView):
    model = Asset
    template_name = 'assets/delete_confirm.html'
    success_url = reverse_lazy('assets:asset-list')


class AssetDetailView(DetailView):
    model = Asset
    context_object_name = 'asset'
    template_name = 'assets/asset_detail.html'

    def get_context_data(self, **kwargs):
        asset_groups = self.object.groups.all()
        system_users = self.object.system_users.all()
        context = {
            'app': 'Assets',
            'action': 'Asset detail',
            'asset_groups_remain': [asset_group for asset_group in AssetGroup.objects.all()
                                    if asset_group not in asset_groups],
            'asset_groups': asset_groups,
            'system_users_remain': [system_user for system_user in SystemUser.objects.all()
                                    if system_user not in system_users],
            'system_users': system_users,
        }
        kwargs.update(context)
        return super(AssetDetailView, self).get_context_data(**kwargs)


class AssetModalListView(AdminUserRequiredMixin, ListView):
    paginate_by = settings.CONFIG.DISPLAY_PER_PAGE
    model = Asset
    context_object_name = 'asset_modal_list'
    template_name = 'assets/asset_modal_list.html'

    def get_context_data(self, **kwargs):
        group_id = self.request.GET.get('group_id')
        plain_id_lists = self.request.GET.get('plain_id_lists')
        self.s = self.request.GET.get('plain_id_lists')
        assets = Asset.objects.all()
        if "," in str(self.s):
            self.plain_id_lists = [int(x) for x in self.s.split(',')]
        else:
            self.plain_id_lists = [self.s]

        if plain_id_lists:
            if "," in str(self.s):
                plain_id_lists = [int(x) for x in self.s.split(',')]
            else:
                plain_id_lists = [int(self.s)]
            context = {
                'all_assets': plain_id_lists,
            }
            kwargs.update(context)
        if group_id:
            group = AssetGroup.objects.get(id=group_id)
            context = {
                'all_assets': [x.id for x in group.assets.all()],
                'assets': assets
            }
            kwargs.update(context)
        else:
            context = {
                'assets': assets
            }
            kwargs.update(context)
        return super(AssetModalListView, self).get_context_data(**kwargs)


@method_decorator(csrf_exempt, name='dispatch')
class AssetExportView(View):
    @staticmethod
    def get_asset_attr(asset, attr):
        if attr in ['admin_user', 'idc']:
            return getattr(asset, attr).name
        elif attr in ['status', 'type', 'env']:
            return getattr(asset, 'get_{}_display'.format(attr))()
        else:
            return getattr(asset, attr)

    def get(self, request, *args, **kwargs):
        spm = request.GET.get('spm', '')
        assets_id = cache.get(spm)
        if not assets_id and not isinstance(assets_id, list):
            return HttpResponse('May be expired', status=404)

        assets = Asset.objects.filter(id__in=assets_id)
        wb = Workbook()
        ws = wb.active
        ws.title = 'Asset'
        header = ['hostname', 'ip', 'port', 'admin_user', 'idc', 'cpu', 'memory', 'disk',
                  'mac_address', 'other_ip', 'remote_card_ip', 'os', 'cabinet_no',
                  'cabinet_pos', 'number', 'status', 'type', 'env', 'sn', 'comment']
        ws.append(header)

        for asset in assets:
            ws.append([self.get_asset_attr(asset, attr) for attr in header])

        filename = 'assets-{}.xlsx'.format(timezone.localtime(timezone.now()).strftime('%Y-%m-%d_%H-%M-%S'))
        response = HttpResponse(save_virtual_workbook(wb), content_type='applications/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        return response

    def post(self, request, *args, **kwargs):
        try:
            assets_id = json.loads(request.body).get('assets_id', [])
            print(assets_id)
        except ValueError:
            return HttpResponse('Json object not valid', status=400)
        spm = uuid.uuid4().get_hex()
        cache.set(spm, assets_id, 300)
        url = reverse_lazy('assets:asset-export') + '?spm=%s' % spm
        return JsonResponse({'redirect': url})


class BulkImportAssetView(AdminUserRequiredMixin, JSONResponseMixin, FormView):
    form_class = forms.FileForm

    def form_valid(self, form):
        try:
            wb = load_workbook(form.cleaned_data['file'])
            ws = wb.get_active_sheet()
        except Exception as e:
            print(e)
            data = {'valid': False, 'msg': 'Not a valid Excel file'}
            return self.render_json_response(data)

        rows = ws.rows
        header_all = ['hostname', 'ip', 'port', 'admin_user', 'idc', 'cpu', 'memory', 'disk',
                      'mac_address', 'other_ip', 'remote_card_ip', 'os', 'cabinet_no',
                      'cabinet_pos', 'number', 'status', 'type', 'env', 'sn', 'comment']
        header_min = ['hostname', 'ip', 'port', 'admin_user', 'comment']
        header = [col.value for col in next(rows)]
        if not set(header).issubset(set(header_all)) and not set(header).issuperset(set(header_min)):
            data = {'valid': False, 'msg': 'Must be same format as template or export file'}
            return self.render_json_response(data)

        created = []
        updated = []
        failed = []
        for row in rows:
            asset_dict = dict(zip(header, [col.value for col in row]))
            if asset_dict.get('admin_user', None):
                admin_user = get_object_or_none(AdminUser, name=asset_dict['admin_user'])
                asset_dict['admin_user'] = admin_user

            if asset_dict.get('idc'):
                idc = get_object_or_none(IDC, name=asset_dict['idc'])
                asset_dict['idc'] = idc

            if asset_dict.get('type'):
                asset_display_type_map = dict(zip(dict(Asset.TYPE_CHOICES).values(), dict(Asset.TYPE_CHOICES).keys()))
                asset_type = asset_display_type_map.get(asset_dict['type'], 'Server')
                asset_dict['type'] = asset_type

            if asset_dict.get('status'):
                asset_display_status_map = dict(zip(dict(Asset.STATUS_CHOICES).values(),
                                                    dict(Asset.STATUS_CHOICES).keys()))
                asset_status = asset_display_status_map.get(asset_dict['status'], 'In use')
                asset_dict['status'] = asset_status

            if asset_dict.get('env'):
                asset_display_env_map = dict(zip(dict(Asset.ENV_CHOICES).values(),
                                                 dict(Asset.ENV_CHOICES).keys()))
                asset_env = asset_display_env_map.get(asset_dict['env'], 'Prod')
                asset_dict['env'] = asset_env

            try:
                Asset.objects.create(**asset_dict)
                created.append(asset_dict['ip'])
            except IntegrityError as e:
                asset = Asset.objects.filter(ip=asset_dict['ip'], port=asset_dict['port'])
                if not asset:
                    failed.append(asset_dict['ip'])
                    continue
                asset.update(**asset_dict)
                updated.append(asset_dict['ip'])
            except TypeError as e:
                print(e)
                failed.append(asset_dict['ip'])

        data = {
            'created': created,
            'created_info': 'Created {}'.format(len(created)),
            'updated': updated,
            'updated_info': 'Updated {}'.format(len(updated)),
            'failed': failed,
            'failed_info': 'Failed {}'.format(len(failed)),
            'valid': True,
            'msg': 'Created: {}. Updated: {}, Error: {}'.format(len(created), len(updated), len(failed))
        }
        return self.render_json_response(data)


class AssetRefreshHardwareView(AdminUserRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        asset_id = kwargs.get('pk')
        asset = get_object_or_404(Asset, pk=asset_id)
        update_assets_hardware_info([asset])

        return redirect(reverse('assets:asset-detail', kwargs={'pk': asset_id}))

