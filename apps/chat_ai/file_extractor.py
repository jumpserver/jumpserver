import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree

import chardet
from openpyxl import load_workbook
from pypdf import PdfReader


PLAIN_TEXT_EXTENSIONS = {
    '.bash', '.bat', '.c', '.cc', '.cfg', '.cmd', '.conf', '.cpp', '.css',
    '.csv', '.cxx', '.env', '.fish', '.go', '.h', '.hpp', '.htm', '.html',
    '.ini', '.java', '.js', '.json', '.jsonl', '.jsx', '.kt', '.kts', '.less',
    '.log', '.lua', '.markdown', '.md', '.mjs', '.php', '.pl', '.properties',
    '.ps1', '.py', '.r', '.rb', '.rs', '.scala', '.scss', '.sh', '.sql',
    '.swift', '.toml', '.ts', '.tsv', '.tsx', '.txt', '.vue', '.xml', '.yaml',
    '.yml', '.zsh',
}
OFFICE_EXTENSIONS = {'.docx', '.pptx', '.xlsx'}
SUPPORTED_FILE_EXTENSIONS = PLAIN_TEXT_EXTENSIONS | OFFICE_EXTENSIONS | {'.pdf'}
SUPPORTED_EXTENSION_LABEL = 'PDF, DOCX, XLSX, PPTX, or a text/code file'

MAX_ZIP_ENTRIES = 2000
MAX_ZIP_UNCOMPRESSED_SIZE = 50 * 1024 * 1024


class FileExtractionError(ValueError):
    pass


def get_file_extension(name):
    path = Path(name)
    if path.name.lower() == '.env':
        return '.env'
    return path.suffix.lower()


def extract_file_text(uploaded, *, max_chars):
    extension = get_file_extension(uploaded.name)
    if extension not in SUPPORTED_FILE_EXTENSIONS:
        raise FileExtractionError(f'Unsupported file format. Use {SUPPORTED_EXTENSION_LABEL}.')

    try:
        uploaded.seek(0)
        if extension in PLAIN_TEXT_EXTENSIONS:
            text = _extract_plain_text(uploaded)
        elif extension == '.pdf':
            text = _extract_pdf(uploaded, max_chars)
        elif extension == '.docx':
            text = _extract_docx(uploaded)
        elif extension == '.pptx':
            text = _extract_pptx(uploaded)
        else:
            text = _extract_xlsx(uploaded, max_chars)
    except FileExtractionError:
        raise
    except Exception as exc:
        raise FileExtractionError(f'File {uploaded.name} could not be read.') from exc
    finally:
        uploaded.seek(0)

    text = _normalize_text(text)
    if not text:
        raise FileExtractionError(
            f'File {uploaded.name} does not contain readable text. Scanned documents are not supported.'
        )
    if len(text) > max_chars:
        text = text[:max_chars].rstrip() + '\n\n[File content truncated]'
    return text


def _extract_plain_text(uploaded):
    data = uploaded.read()
    if b'\x00' in data[:4096]:
        raise FileExtractionError(f'File {uploaded.name} appears to be binary.')
    try:
        text = data.decode('utf-8-sig')
    except UnicodeDecodeError:
        encoding = (chardet.detect(data).get('encoding') or '').lower()
        if not encoding:
            raise FileExtractionError(f'File {uploaded.name} has an unsupported text encoding.')
        try:
            text = data.decode(encoding)
        except (LookupError, UnicodeDecodeError) as exc:
            raise FileExtractionError(
                f'File {uploaded.name} has an unsupported text encoding.'
            ) from exc
    control_chars = sum(
        1 for char in text
        if ord(char) < 32 and char not in '\n\r\t\f'
    )
    if text and control_chars / len(text) > 0.02:
        raise FileExtractionError(f'File {uploaded.name} appears to be binary.')
    return text


def _extract_pdf(uploaded, max_chars):
    reader = PdfReader(uploaded, strict=False)
    if reader.is_encrypted and not reader.decrypt(''):
        raise FileExtractionError('Password-protected PDF files are not supported.')
    parts = []
    length = 0
    for index, page in enumerate(reader.pages[:200], start=1):
        page_text = page.extract_text() or ''
        if page_text.strip():
            part = f'[Page {index}]\n{page_text}'
            parts.append(part)
            length += len(part)
        if length >= max_chars:
            break
    return '\n\n'.join(parts)


def _open_office_archive(uploaded):
    archive = zipfile.ZipFile(uploaded)
    infos = archive.infolist()
    if len(infos) > MAX_ZIP_ENTRIES:
        archive.close()
        raise FileExtractionError(f'File {uploaded.name} contains too many entries.')
    if sum(info.file_size for info in infos) > MAX_ZIP_UNCOMPRESSED_SIZE:
        archive.close()
        raise FileExtractionError(f'File {uploaded.name} expands beyond the configured limit.')
    return archive


def _xml_text(data):
    root = ElementTree.fromstring(data)
    return ' '.join(
        element.text.strip()
        for element in root.iter()
        if element.tag.rsplit('}', 1)[-1] == 't' and element.text and element.text.strip()
    )


def _extract_docx(uploaded):
    with _open_office_archive(uploaded) as archive:
        if 'word/document.xml' not in archive.namelist():
            raise FileExtractionError(f'File {uploaded.name} is not a valid DOCX document.')
        return _xml_text(archive.read('word/document.xml'))


def _natural_key(value):
    return [int(part) if part.isdigit() else part for part in re.split(r'(\d+)', value)]


def _extract_pptx(uploaded):
    with _open_office_archive(uploaded) as archive:
        slides = sorted(
            (
                name for name in archive.namelist()
                if re.fullmatch(r'ppt/slides/slide\d+\.xml', name)
            ),
            key=_natural_key,
        )
        if not slides:
            raise FileExtractionError(f'File {uploaded.name} is not a valid PPTX presentation.')
        return '\n\n'.join(
            f'[Slide {index}]\n{_xml_text(archive.read(name))}'
            for index, name in enumerate(slides[:200], start=1)
        )


def _extract_xlsx(uploaded, max_chars):
    with _open_office_archive(uploaded):
        pass
    uploaded.seek(0)
    workbook = load_workbook(uploaded, read_only=True, data_only=True)
    try:
        parts = []
        length = 0
        for sheet in workbook.worksheets:
            parts.append(f'[Sheet: {sheet.title}]')
            for row in sheet.iter_rows(values_only=True):
                values = ['' if value is None else str(value) for value in row]
                if not any(values):
                    continue
                line = '\t'.join(values).rstrip()
                parts.append(line)
                length += len(line) + 1
                if length >= max_chars:
                    return '\n'.join(parts)
        return '\n'.join(parts)
    finally:
        workbook.close()


def _normalize_text(value):
    value = value.replace('\r\n', '\n').replace('\r', '\n').replace('\x00', '')
    value = re.sub(r'[ \t]+\n', '\n', value)
    value = re.sub(r'\n{4,}', '\n\n\n', value)
    return value.strip()
