# ~*~ coding: utf-8 ~*~

from __future__ import unicode_literals

import json
import uuid

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.core.cache import cache
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect
from django.urls import reverse_lazy, reverse
from django.utils import timezone
from django.utils.translation import ugettext as _
from django.utils.decorators import method_decorator
from django.views import View
from django.views.generic import ListView
from django.views.generic.base import TemplateView
from django.views.generic.edit import (CreateView, UpdateView, FormMixin,
                                       FormView)
from django.views.generic.detail import DetailView, SingleObjectMixin
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import logout as auth_logout

from .. import forms
from ..models import User, UserGroup
from ..utils import AdminUserRequiredMixin, user_add_success_next
from common.mixins import JSONResponseMixin
from common.utils import get_logger
from perms.models import AssetPermission

__all__ = ['UserListView', 'UserCreateView', 'UserDetailView',
           'UserUpdateView', 'UserAssetPermissionCreateView',
           'UserAssetPermissionView', 'UserGrantedAssetView',
           'UserExportView',  'UserBulkImportView', 'UserProfileView',
           'UserProfileUpdateView', 'UserPasswordUpdateView',
           'UserPublicKeyUpdateView',
           ]

logger = get_logger(__name__)


class UserListView(AdminUserRequiredMixin, TemplateView):
    template_name = 'users/user_list.html'

    def get_context_data(self, **kwargs):
        context = super(UserListView, self).get_context_data(**kwargs)
        context.update({
            'app': _('Users'),
            'action': _('User list'),
            'groups': UserGroup.objects.all()
        })
        return context


class UserCreateView(AdminUserRequiredMixin, SuccessMessageMixin, CreateView):
    model = User
    form_class = forms.UserCreateUpdateForm
    template_name = 'users/user_create.html'
    success_url = reverse_lazy('users:user-list')
    success_message = _('Create user <a href="{url}">{name}</a> successfully.')

    def get_context_data(self, **kwargs):
        context = super(UserCreateView, self).get_context_data(**kwargs)
        context.update({'app': _('Users'), 'action': _('Create user')})
        return context

    def form_valid(self, form):
        user = form.save(commit=False)
        user.created_by = self.request.user.username or 'System'
        user.save()
        user_add_success_next(user)
        return super(UserCreateView, self).form_valid(form)

    def get_success_message(self, cleaned_data):
        url = reverse_lazy('users:user-detail', kwargs={'pk': self.object.pk})
        return self.success_message.format(
            url=url, name=self.object.name
        )


class UserUpdateView(AdminUserRequiredMixin, UpdateView):
    model = User
    form_class = forms.UserCreateUpdateForm
    template_name = 'users/user_update.html'
    context_object_name = 'user_object'
    success_url = reverse_lazy('users:user-list')

    def form_valid(self, form):
        username = self.object.username
        user = form.save(commit=False)
        user.username = username
        user.save()
        password = self.request.POST.get('password', '')
        if password:
            user.set_password(password)
        return super(UserUpdateView, self).form_valid(form)

    def get_context_data(self, **kwargs):
        context = super(UserUpdateView, self).get_context_data(**kwargs)
        context.update({'app': _('Users'), 'action': _('Update user')})
        return context


class UserDetailView(AdminUserRequiredMixin, DetailView):
    model = User
    template_name = 'users/user_detail.html'
    context_object_name = "user_object"

    def get_context_data(self, **kwargs):
        groups = UserGroup.objects.exclude(id__in=self.object.groups.all())
        context = {
            'app': _('Users'),
            'action': _('User detail'),
            'groups': groups
        }
        kwargs.update(context)
        return super(UserDetailView, self).get_context_data(**kwargs)


@method_decorator(csrf_exempt, name='dispatch')
class UserExportView(View):
    def get(self, request):
        spm = request.GET.get('spm', '')
        users_id = cache.get(spm)
        if not users_id and not isinstance(users_id, list):
            return HttpResponse('May be expired', status=404)

        users = User.objects.filter(id__in=users_id)
        print(users)
        wb = Workbook()
        ws = wb.active
        ws.title = 'User'
        header = ["name", 'username', 'email', 'groups',
                  "role", "phone", "wechat", "comment"]
        ws.append(header)

        for user in users:
            ws.append([user.name, user.username, user.email,
                       ','.join([group.name for group in user.groups.all()]),
                       user.role, user.phone, user.wechat, user.comment])

        filename = 'users-{}.xlsx'.format(
            timezone.localtime(timezone.now()).strftime('%Y-%m-%d_%H-%M-%S'))
        response = HttpResponse(save_virtual_workbook(wb),
                                content_type='applications/vnd.ms-excel')
        response[
            'Content-Disposition'] = 'attachment; filename="%s"' % filename
        return response

    def post(self, request):
        try:
            users_id = json.loads(request.body).get('users_id', [])
        except ValueError:
            return HttpResponse('Json object not valid', status=400)
        spm = uuid.uuid4().get_hex()
        cache.set(spm, users_id, 300)
        url = reverse('users:user-export') + '?spm=%s' % spm
        return JsonResponse({'redirect': url})


class UserBulkImportView(AdminUserRequiredMixin, JSONResponseMixin, FormView):
    form_class = forms.FileForm

    def form_invalid(self, form):
        try:
            error = form.errors.values()[-1][-1]
        except Exception as e:
            error = _('Invalid file.')
        data = {
            'success': False,
            'msg': error
        }
        return self.render_json_response(data)

    def form_valid(self, form):
        try:
            wb = load_workbook(form.cleaned_data['file'])
            ws = wb.get_active_sheet()
        except Exception as e:
            print(e)
            data = {'valid': False, 'msg': 'Not a valid Excel file'}
            return self.render_json_response(data)

        rows = ws.rows
        header_need = ["name", 'username', 'email', 'groups',
                       "role", "phone", "wechat", "comment"]
        header = [col.value for col in next(rows)]
        print(header)
        if header != header_need:
            data = {'valid': False, 'msg': 'Must be same format as '
                                           'template or export file'}
            return self.render_json_response(data)

        created = []
        updated = []
        failed = []
        for row in rows:
            user_dict = dict(zip(header, [col.value for col in row]))
            groups_name = user_dict.pop('groups')
            if groups_name:
                groups_name = groups_name.split(',')
                groups = UserGroup.objects.filter(name__in=groups_name)
            else:
                groups = None
            try:
                user = User.objects.create(**user_dict)
                user_add_success_next(user)
                created.append(user_dict['username'])
            except User.IntegrityError as e:
                user = User.objects.filter(username=user_dict['username'])
                if not user:
                    failed.append(user_dict['username'])
                    continue
                user.update(**user_dict)
                user = user[0]
                updated.append(user_dict['username'])
            except TypeError as e:
                print(e)
                failed.append(user_dict['username'])
                user = None

            if user and groups:
                user.groups.add(*tuple(groups))
                user.save()

        data = {
            'created': created,
            'created_info': 'Created {}'.format(len(created)),
            'updated': updated,
            'updated_info': 'Updated {}'.format(len(updated)),
            'failed': failed,
            'failed_info': 'Failed {}'.format(len(failed)),
            'valid': True,
            'msg': 'Created: {}. Updated: {}, Error: {}'.format(
                len(created), len(updated), len(failed))
        }
        return self.render_json_response(data)


class UserAssetPermissionView(AdminUserRequiredMixin, FormMixin,
                              SingleObjectMixin, ListView):
    model = User
    template_name = 'users/user_asset_permission.html'
    context_object_name = 'user'
    form_class = forms.UserPrivateAssetPermissionForm

    def get(self, request, *args, **kwargs):
        self.object = self.get_object(queryset=User.objects.all())
        return super(UserAssetPermissionView, self).get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = {
            'app': 'Users',
            'action': 'User asset permissions',
        }
        kwargs.update(context)
        return super(UserAssetPermissionView, self).get_context_data(**kwargs)


class UserAssetPermissionCreateView(AdminUserRequiredMixin, CreateView):
    form_class = forms.UserPrivateAssetPermissionForm
    model = AssetPermission

    def get(self, request, *args, **kwargs):
        user = self.get_object(queryset=User.objects.all())
        return redirect(reverse('users:user-asset-permission',
                                kwargs={'pk': user.id}))

    def post(self, request, *args, **kwargs):
        self.user = self.get_object(queryset=User.objects.all())
        return super(UserAssetPermissionCreateView, self)\
            .post(request, *args, **kwargs)

    def get_form(self, form_class=None):
        form = super(UserAssetPermissionCreateView, self)\
            .get_form(form_class=form_class)
        form.user = self.user
        return form

    def form_invalid(self, form):
        return redirect(reverse('users:user-asset-permission',
                                kwargs={'pk': self.user.id}))

    def get_success_url(self):
        return reverse('users:user-asset-permission',
                       kwargs={'pk': self.user.id})


class UserGrantedAssetView(AdminUserRequiredMixin, DetailView):
    model = User
    template_name = 'users/user_granted_asset.html'
    context_object_name = 'user'

    def get(self, request, *args, **kwargs):
        self.object = self.get_object(queryset=User.objects.all())
        return super(UserGrantedAssetView, self).get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = {
            'app': 'User',
            'action': 'User granted asset',
        }
        kwargs.update(context)
        return super(UserGrantedAssetView, self).get_context_data(**kwargs)


class UserProfileView(LoginRequiredMixin, TemplateView):
    template_name = 'users/user_profile.html'

    def get_context_data(self, **kwargs):
        from perms.utils import get_user_granted_assets
        assets = get_user_granted_assets(self.request.user)
        context = {
            'app': 'User',
            'action': 'User Profile',
            'assets': assets,
        }
        kwargs.update(context)
        return super(UserProfileView, self).get_context_data(**kwargs)


class UserProfileUpdateView(LoginRequiredMixin, UpdateView):
    template_name = 'users/user_profile_update.html'
    model = User
    form_class = forms.UserProfileForm
    success_url = reverse_lazy('users:user-profile')
    success_message = _('Create user <a href="{url}">{name}</a> successfully.')

    def get_object(self, queryset=None):
        return self.request.user

    def get_success_message(self, cleaned_data):
        url = reverse_lazy('users:user-detail', kwargs={'pk': self.object.pk})
        return self.success_message.format(
            url=url, name=self.object.name
        )

    def get_context_data(self, **kwargs):
        context = {
            'app': 'User',
            'action': 'Profile update',
        }
        kwargs.update(context)
        return super(UserProfileUpdateView, self).get_context_data(**kwargs)


class UserPasswordUpdateView(LoginRequiredMixin, UpdateView):
    template_name = 'users/user_password_update.html'
    model = User
    form_class = forms.UserPasswordForm
    success_url = reverse_lazy('users:user-profile')

    def get_object(self, queryset=None):
        return self.request.user

    def get_context_data(self, **kwargs):
        context = {
            'app': 'User',
            'action': 'Password update',
        }
        kwargs.update(context)
        return super(UserPasswordUpdateView, self).get_context_data(**kwargs)

    def get_success_url(self):
        auth_logout(self.request)
        return super(UserPasswordUpdateView, self).get_success_url()


class UserPublicKeyUpdateView(LoginRequiredMixin, UpdateView):
    template_name = 'users/user_pubkey_update.html'
    model = User
    form_class = forms.UserPublicKeyForm
    success_url = reverse_lazy('users:user-profile')

    def get_object(self, queryset=None):
        return self.request.user

    def get_context_data(self, **kwargs):
        context = {
            'app': 'User',
            'action': 'Public key update',
        }
        kwargs.update(context)
        return super(UserPublicKeyUpdateView, self).get_context_data(**kwargs)
