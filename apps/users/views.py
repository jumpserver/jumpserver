# ~*~ coding: utf-8 ~*~

from __future__ import unicode_literals
import json
import uuid
import codecs

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import load_workbook
import unicodecsv as csv
from django import forms
from django.utils import timezone
from django.core.cache import cache
from django.db import IntegrityError
from django.contrib.auth import login as auth_login, logout as auth_logout
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.core.files.storage import default_storage
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.shortcuts import reverse, redirect
from django.utils.decorators import method_decorator
from django.utils.translation import ugettext as _
from django.urls import reverse_lazy
from django.views import View
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.views.decorators.debug import sensitive_post_parameters
from django.views.generic.base import TemplateView
from django.views.generic.list import ListView
from django.views.generic.edit import CreateView, DeleteView, UpdateView, FormView, SingleObjectMixin, FormMixin
from django.views.generic.detail import DetailView
from formtools.wizard.views import SessionWizardView

from common.mixins import JSONResponseMixin
from common.utils import get_object_or_none, get_logger
from perms.models import AssetPermission
from .models import User, UserGroup
from .utils import AdminUserRequiredMixin, user_add_success_next, send_reset_password_mail
from .hands import write_login_log_async
from . import forms

logger = get_logger(__name__)


@method_decorator(sensitive_post_parameters(), name='dispatch')
@method_decorator(csrf_protect, name='dispatch')
@method_decorator(never_cache, name='dispatch')
class UserLoginView(FormView):
    template_name = 'users/login.html'
    form_class = forms.UserLoginForm
    redirect_field_name = 'next'

    def get(self, request, *args, **kwargs):
        if request.user.is_staff:
            return redirect(self.get_success_url())
        return super(UserLoginView, self).get(request, *args, **kwargs)

    def form_valid(self, form):
        auth_login(self.request, form.get_user())
        login_ip = self.request.META.get('REMOTE_ADDR', '')
        user_agent = self.request.META.get('HTTP_USER_AGENT', '')
        write_login_log_async.delay(self.request.user.username, self.request.user.name,
                                    login_type='W', login_ip=login_ip, user_agent=user_agent)
        return redirect(self.get_success_url())

    def get_success_url(self):
        if self.request.user.is_first_login:
            return reverse('users:user-first-login')

        return self.request.POST.get(
            self.redirect_field_name,
            self.request.GET.get(self.redirect_field_name, reverse('index')))


@method_decorator(never_cache, name='dispatch')
class UserLogoutView(TemplateView):
    template_name = 'flash_message_standalone.html'

    def get(self, request, *args, **kwargs):
        auth_logout(request)
        return super(UserLogoutView, self).get(request)

    def get_context_data(self, **kwargs):
        context = {
            'title': _('Logout success'),
            'messages': _('Logout success, return login page'),
            'redirect_url': reverse('users:login'),
            'auto_redirect': True,
        }
        kwargs.update(context)
        return super(UserLogoutView, self).get_context_data(**kwargs)


class UserListView(AdminUserRequiredMixin, TemplateView):
    template_name = 'users/user_list.html'

    def get_context_data(self, **kwargs):
        context = super(UserListView, self).get_context_data(**kwargs)
        context.update({
            'app': _('Users'),
            'action': _('User list'),
            'groups': UserGroup.objects.all()
        })
        return context


class UserCreateView(AdminUserRequiredMixin, SuccessMessageMixin, CreateView):
    model = User
    form_class = forms.UserCreateUpdateForm
    template_name = 'users/user_create.html'
    success_url = reverse_lazy('users:user-list')
    success_message = _('Create user <a href="%s">%s</a> successfully.')

    def get_context_data(self, **kwargs):
        context = super(UserCreateView, self).get_context_data(**kwargs)
        context.update({'app': _('Users'), 'action': _('Create user')})
        return context

    def form_valid(self, form):
        user = form.save(commit=False)
        user.created_by = self.request.user.username or 'System'
        user.save()
        user_add_success_next(user)
        return super(UserCreateView, self).form_valid(form)

    def get_success_message(self, cleaned_data):
        return self.success_message % (
            reverse_lazy('users:user-detail', kwargs={'pk': self.object.pk}),
            self.object.name,
        )


class UserUpdateView(AdminUserRequiredMixin, UpdateView):
    model = User
    form_class = forms.UserCreateUpdateForm
    template_name = 'users/user_update.html'
    context_object_name = 'user_object'
    success_url = reverse_lazy('users:user-list')

    def form_valid(self, form):
        username = self.object.username
        user = form.save(commit=False)
        user.username = username
        user.save()
        password = self.request.POST.get('password', '')
        if password:
            user.set_password(password)
        return super(UserUpdateView, self).form_valid(form)

    def get_context_data(self, **kwargs):
        context = super(UserUpdateView, self).get_context_data(**kwargs)
        context.update({'app': _('Users'), 'action': _('Update user')})
        return context


class UserDetailView(AdminUserRequiredMixin, DetailView):
    model = User
    template_name = 'users/user_detail.html'
    context_object_name = "user"

    def get_context_data(self, **kwargs):
        groups = UserGroup.objects.exclude(id__in=self.object.groups.all())
        context = {'app': _('Users'), 'action': _('User detail'), 'groups': groups}
        kwargs.update(context)
        return super(UserDetailView, self).get_context_data(**kwargs)


class UserGroupListView(AdminUserRequiredMixin, TemplateView):
    template_name = 'users/user_group_list.html'

    def get_context_data(self, **kwargs):
        context = super(UserGroupListView, self).get_context_data(**kwargs)
        context.update({'app': _('Users'), 'action': _('User group list')})
        return context


class UserGroupCreateView(AdminUserRequiredMixin, CreateView):
    model = UserGroup
    form_class = forms.UserGroupForm
    template_name = 'users/user_group_create_update.html'
    success_url = reverse_lazy('users:user-group-list')

    def get_context_data(self, **kwargs):
        context = super(UserGroupCreateView, self).get_context_data(**kwargs)
        users = User.objects.all()
        context.update({'app': _('Users'), 'action': _('Create user group'), 'users': users})
        return context

    def form_valid(self, form):
        user_group = form.save()
        users_id_list = self.request.POST.getlist('users', [])
        users = User.objects.filter(id__in=users_id_list)
        user_group.created_by = self.request.user.username or 'Admin'
        user_group.users.add(*users)
        user_group.save()
        return super(UserGroupCreateView, self).form_valid(form)


class UserGroupUpdateView(AdminUserRequiredMixin, UpdateView):
    model = UserGroup
    form_class = forms.UserGroupForm
    template_name = 'users/user_group_create_update.html'
    success_url = reverse_lazy('users:user-group-list')

    def get_context_data(self, **kwargs):
        # self.object = self.get_object()
        context = super(UserGroupUpdateView, self).get_context_data(**kwargs)
        users = User.objects.all()
        group_users = [user.id for user in self.object.users.all()]
        context.update({
            'app': _('Users'),
            'action': _('Update User Group'),
            'users': users,
            'group_users': group_users
        })
        return context

    def form_valid(self, form):
        user_group = form.save()
        users_id_list = self.request.POST.getlist('users', [])
        users = User.objects.filter(id__in=users_id_list)
        user_group.users.clear()
        user_group.users.add(*users)
        user_group.save()
        return super(UserGroupUpdateView, self).form_valid(form)


class UserGroupDetailView(AdminUserRequiredMixin, DetailView):
    model = UserGroup
    context_object_name = 'user_group'
    template_name = 'users/user_group_detail.html'

    def get_context_data(self, **kwargs):
        users = User.objects.exclude(id__in=self.object.users.all())
        context = {
            'app': _('Users'),
            'action': _('User Group Detail'),
            'users': users,
        }
        kwargs.update(context)
        return super(UserGroupDetailView, self).get_context_data(**kwargs)


class UserGroupDeleteView(DeleteView):
    pass


class UserForgotPasswordView(TemplateView):
    template_name = 'users/forgot_password.html'

    def post(self, request):
        email = request.POST.get('email')
        user = get_object_or_none(User, email=email)
        if not user:
            return self.get(request, errors=_('Email address invalid, input again'))
        else:
            send_reset_password_mail(user)
            return HttpResponseRedirect(reverse('users:forgot-password-sendmail-success'))


class UserForgotPasswordSendmailSuccessView(TemplateView):
    template_name = 'flash_message_standalone.html'

    def get_context_data(self, **kwargs):
        context = {
            'title': _('Send reset password message'),
            'messages': _('Send reset password mail success, login your mail box and follow it '),
            'redirect_url': reverse('users:login'),
        }
        kwargs.update(context)
        return super(UserForgotPasswordSendmailSuccessView, self).get_context_data(**kwargs)


class UserResetPasswordSuccessView(TemplateView):
    template_name = 'flash_message_standalone.html'

    def get_context_data(self, **kwargs):
        context = {
            'title': _('Reset password success'),
            'messages': _('Reset password success, return to login page'),
            'redirect_url': reverse('users:login'),
            'auto_redirect': True,
        }
        kwargs.update(context)
        return super(UserResetPasswordSuccessView, self).get_context_data(**kwargs)


class UserResetPasswordView(TemplateView):
    template_name = 'users/reset_password.html'

    def get(self, request, *args, **kwargs):
        token = request.GET.get('token')
        user = User.validate_reset_token(token)

        if not user:
            kwargs.update({'errors': _('Token invalid or expired')})
        return super(UserResetPasswordView, self).get(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        password = request.POST.get('password')
        password_confirm = request.POST.get('password-confirm')
        token = request.GET.get('token')

        if password != password_confirm:
            return self.get(request, errors=_('Password not same'))

        user = User.validate_reset_token(token)
        if not user:
            return self.get(request, errors=_('Token invalid or expired'))

        user.reset_password(password)
        return HttpResponseRedirect(reverse('users:reset-password-success'))


class UserFirstLoginView(LoginRequiredMixin, SessionWizardView):
    template_name = 'users/first_login.html'
    form_list = [forms.UserInfoForm, forms.UserKeyForm]
    file_storage = default_storage

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated() and not request.user.is_first_login:
            return redirect(reverse('index'))
        return super(UserFirstLoginView, self).dispatch(request, *args, **kwargs)

    def done(self, form_list, form_dict, **kwargs):
        user = self.request.user
        for form in form_list:
            for field in form:
                if field.value():
                    setattr(user, field.name, field.value())
                if field.name == 'enable_otp':
                    user.enable_otp = field.value()
        user.is_first_login = False
        user.is_public_key_valid = True
        user.save()
        return redirect(reverse('index'))

    def get_context_data(self, **kwargs):
        context = super(UserFirstLoginView, self).get_context_data(**kwargs)
        context.update({'app': _('Users'), 'action': _('First Login')})
        return context

    def get_form_initial(self, step):
        user = self.request.user
        if step == '0':
            return {
                'name': user.name or user.username,
                'enable_otp': user.enable_otp or True,
                'wechat': user.wechat or '',
                'phone': user.phone or ''
            }
        return super(UserFirstLoginView, self).get_form_initial(step)

    def get_form(self, step=None, data=None, files=None):
        form = super(UserFirstLoginView, self).get_form(step, data, files)

        if step is None:
            step = self.steps.current

        if step == '1':
            form.user = self.request.user
        return form


class UserAssetPermissionView(AdminUserRequiredMixin, FormMixin, SingleObjectMixin, ListView):
    model = User
    template_name = 'users/user_asset_permission.html'
    context_object_name = 'user'
    form_class = forms.UserPrivateAssetPermissionForm

    def get(self, request, *args, **kwargs):
        self.object = self.get_object(queryset=User.objects.all())
        return super(UserAssetPermissionView, self).get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = {
            'app': 'Users',
            'action': 'User asset permissions',
        }
        kwargs.update(context)
        return super(UserAssetPermissionView, self).get_context_data(**kwargs)


class UserGroupAssetPermissionView(AdminUserRequiredMixin, FormMixin, SingleObjectMixin, ListView):
    model = UserGroup
    template_name = 'users/user_group_asset_permission.html'
    context_object_name = 'user_group'
    form_class = forms.UserPrivateAssetPermissionForm

    def get(self, request, *args, **kwargs):
        self.object = self.get_object(queryset=UserGroup.objects.all())
        return super(UserGroupAssetPermissionView, self).get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = {
            'app': 'Users',
            'action': 'User group asset permissions',
        }
        kwargs.update(context)
        return super(UserGroupAssetPermissionView, self).get_context_data(**kwargs)


class UserAssetPermissionCreateView(AdminUserRequiredMixin, CreateView):
    form_class = forms.UserPrivateAssetPermissionForm
    model = AssetPermission

    def get(self, request, *args, **kwargs):
        user = self.get_object(queryset=User.objects.all())
        return redirect(reverse('users:user-asset-permission', kwargs={'pk': user.id}))

    def post(self, request, *args, **kwargs):
        self.user = self.get_object(queryset=User.objects.all())
        return super(UserAssetPermissionCreateView, self).post(request, *args, **kwargs)

    def get_form(self, form_class=None):
        form = super(UserAssetPermissionCreateView, self).get_form(form_class=form_class)
        form.user = self.user
        return form

    def form_invalid(self, form):
        return redirect(reverse('users:user-asset-permission', kwargs={'pk': self.user.id}))

    def get_success_url(self):
        return reverse('users:user-asset-permission', kwargs={'pk': self.user.id})


class UserGroupAssetPermissionCreateView(AdminUserRequiredMixin, CreateView):
    form_class = forms.UserGroupPrivateAssetPermissionForm
    model = AssetPermission

    def get(self, request, *args, **kwargs):
        user_group = self.get_object(queryset=UserGroup.objects.all())
        return redirect(reverse('users:user-group-asset-permission', kwargs={'pk': user_group.id}))

    def post(self, request, *args, **kwargs):
        self.user_group = self.get_object(queryset=UserGroup.objects.all())
        return super(UserGroupAssetPermissionCreateView, self).post(request, *args, **kwargs)

    def get_form(self, form_class=None):
        form = super(UserGroupAssetPermissionCreateView, self).get_form(form_class=form_class)
        form.user_group = self.user_group
        return form

    def form_invalid(self, form):
        return redirect(reverse('users:user-group-asset-permission', kwargs={'pk': self.user_group.id}))

    def get_success_url(self):
        return reverse('users:user-group-asset-permission', kwargs={'pk': self.user_group.id})


class UserGrantedAssetView(AdminUserRequiredMixin, DetailView):
    model = User
    template_name = 'users/user_granted_asset.html'
    context_object_name = 'user'

    def get(self, request, *args, **kwargs):
        self.object = self.get_object(queryset=User.objects.all())
        return super(UserGrantedAssetView, self).get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = {
            'app': 'User',
            'action': 'User granted asset',
        }
        kwargs.update(context)
        return super(UserGrantedAssetView, self).get_context_data(**kwargs)


class UserGroupGrantedAssetView(AdminUserRequiredMixin, DetailView):
    model = User
    template_name = 'users/user_group_granted_asset.html'
    context_object_name = 'user_group'

    def get(self, request, *args, **kwargs):
        self.object = self.get_object(queryset=UserGroup.objects.all())
        return super(UserGroupGrantedAssetView, self).get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = {
            'app': 'User',
            'action': 'User group granted asset',
        }
        kwargs.update(context)
        return super(UserGroupGrantedAssetView, self).get_context_data(**kwargs)


class BulkImportUserView(AdminUserRequiredMixin, JSONResponseMixin, FormView):
    form_class = forms.FileForm

    def form_invalid(self, form):
        try:
            error = form.errors.values()[-1][-1]
        except Exception as e:
            print e
            error = _('Invalid file.')
        data = {
            'success': False,
            'msg': error
        }
        return self.render_json_response(data)

    def form_valid(self, form):
        try:
            wb = load_workbook(form.cleaned_data['file'])
            ws = wb.get_active_sheet()
        except Exception as e:
            print(e)
            data = {'valid': False, 'msg': 'Not a valid Excel file'}
            return self.render_json_response(data)

        rows = ws.rows
        header_need = ["name", 'username', 'email', 'groups', "role", "phone", "wechat", "comment"]
        header = [col.value for col in next(rows)]
        print(header)
        if header != header_need:
            data = {'valid': False, 'msg': 'Must be same format as template or export file'}
            return self.render_json_response(data)

        created = []
        updated = []
        failed = []
        for row in rows:
            user_dict = dict(zip(header, [col.value for col in row]))
            groups_name = user_dict.pop('groups')
            if groups_name:
                groups_name = groups_name.split(',')
                groups = UserGroup.objects.filter(name__in=groups_name)
            else:
                groups = None
            try:
                user = User.objects.create(**user_dict)
                created.append(user_dict['username'])
            except IntegrityError:
                user = User.objects.filter(username=user_dict['username'])
                if not user:
                    failed.append(user_dict['username'])
                    continue
                user.update(**user_dict)
                user = user[0]
                updated.append(user_dict['username'])
            except TypeError:
                failed.append(user_dict['username'])
                user = None

            if user and groups:
                user.groups.add(*tuple(groups))
                user.save()

        data = {
            'created': created,
            'created_info': 'Created {}'.format(len(created)),
            'updated': updated,
            'updated_info': 'Updated {}'.format(len(updated)),
            'failed': failed,
            'failed_info': 'Failed {}'.format(len(failed)),
            'valid': True,
            'msg': 'Created: {}. Updated: {}, Error: {}'.format(len(created), len(updated), len(failed))
        }
        return self.render_json_response(data)


@method_decorator(csrf_exempt, name='dispatch')
class ExportUserCsvView(View):
    def get(self, request, *args, **kwargs):
        spm = request.GET.get('spm', '')
        users_id = cache.get(spm)
        if not users_id and not isinstance(users_id, list):
            return HttpResponse('May be expired', status=404)

        users = User.objects.filter(id__in=users_id)
        wb = Workbook()
        ws = wb.active
        ws.title = 'User'
        header = ["name", 'username', 'email', 'groups', "role", "phone", "wechat", "comment"]
        ws.append(header)

        for user in users:
            ws.append([user.name, user.username, user.email,
                       ','.join([group.name for group in user.groups.all()]),
                        user.role, user.phone, user.wechat, user.comment])

        filename = 'users-{}.xlsx'.format(timezone.localtime(timezone.now()).strftime('%Y-%m-%d_%H-%M-%S'))
        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="%s"' % filename
        return response

    def post(self, request, *args, **kwargs):
        try:
            users_id = json.loads(request.body).get('users_id', [])
        except ValueError:
            return HttpResponse('Json object not valid', status=400)
        spm = uuid.uuid4().get_hex()
        cache.set(spm, users_id, 300)
        url = reverse('users:export-user-csv') + '?spm=%s' % spm
        return JsonResponse({'redirect': url})

